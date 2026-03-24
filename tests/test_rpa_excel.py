"""rpa 工作表解析单元测试（unittest + 临时 xlsx）。"""

from __future__ import annotations

import logging
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from app.rpa_excel import (
    HEADER_DDL,
    HEADER_STORAGE_PATH,
    HEADER_TABLE_DESCRIPTION,
    HEADER_TABLE_TYPE,
    HEADER_WAREHOUSE_LAYER,
    SHEET_NAME,
    load_rpa_sheet_hive_rows,
)


def _write_rpa_xlsx(
    path: Path,
    headers: list[str],
    data_rows: list[list[object]] | None,
) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = SHEET_NAME
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    if data_rows:
        for r, row_vals in enumerate(data_rows, start=2):
            for c, val in enumerate(row_vals, start=1):
                ws.cell(row=r, column=c, value=val)
    wb.save(path)


class TestLoadRpaSheetHiveRows(unittest.TestCase):
    def test_warehouse_layer_from_cell_uppercases(self) -> None:
        headers = [
            HEADER_TABLE_DESCRIPTION,
            HEADER_DDL,
            HEADER_STORAGE_PATH,
            HEADER_WAREHOUSE_LAYER,
        ]
        data_rows = [["desc", "CREATE TABLE t (i INT);", "/path", "dwd"]]
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "t.xlsx"
            _write_rpa_xlsx(p, headers, data_rows)
            rows = load_rpa_sheet_hive_rows(p)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].warehouse_layer, "DWD")
        self.assertEqual(rows[0].table_description, "desc")
        self.assertIn("CREATE TABLE", rows[0].ddl_sql)

    def test_warehouse_layer_defaults_when_column_absent(self) -> None:
        headers = [HEADER_TABLE_DESCRIPTION, HEADER_DDL, HEADER_STORAGE_PATH]
        data_rows = [["d", "sql", "/p"]]
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "t.xlsx"
            _write_rpa_xlsx(p, headers, data_rows)
            rows = load_rpa_sheet_hive_rows(p)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].warehouse_layer, "ODS")

    def test_warehouse_layer_defaults_when_cell_empty(self) -> None:
        headers = [
            HEADER_TABLE_DESCRIPTION,
            HEADER_DDL,
            HEADER_STORAGE_PATH,
            HEADER_WAREHOUSE_LAYER,
        ]
        data_rows = [["d", "sql", "/p", ""]]
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "t.xlsx"
            _write_rpa_xlsx(p, headers, data_rows)
            rows = load_rpa_sheet_hive_rows(p)
        self.assertEqual(rows[0].warehouse_layer, "ODS")

    def test_raises_when_required_column_missing(self) -> None:
        headers = [HEADER_TABLE_DESCRIPTION, HEADER_STORAGE_PATH]
        data_rows = [["d", "/p"]]
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "t.xlsx"
            _write_rpa_xlsx(p, headers, data_rows)
            with self.assertRaises(ValueError) as ctx:
                load_rpa_sheet_hive_rows(p)
        self.assertIn(HEADER_DDL, str(ctx.exception))

    def test_raises_when_no_effective_data_row(self) -> None:
        headers = [HEADER_TABLE_DESCRIPTION, HEADER_DDL, HEADER_STORAGE_PATH]
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "t.xlsx"
            _write_rpa_xlsx(p, headers, None)
            with self.assertRaises(ValueError) as ctx:
                load_rpa_sheet_hive_rows(p)
        self.assertIn("没有有效数据行", str(ctx.exception))

    def test_raises_file_not_found(self) -> None:
        with self.assertRaises(FileNotFoundError):
            load_rpa_sheet_hive_rows(Path("/nonexistent/dir/no.xlsx"))

    def test_hive_then_clickhouse_returns_one_hive_row(self) -> None:
        headers = [
            HEADER_TABLE_DESCRIPTION,
            HEADER_DDL,
            HEADER_STORAGE_PATH,
            HEADER_WAREHOUSE_LAYER,
            HEADER_TABLE_TYPE,
        ]
        data_rows = [
            ["h1", "CREATE TABLE a (i INT);", "/a", "ods", "hive"],
            ["", "CREATE TABLE ck (i Int32);", "", "ods", "clickhouse"],
        ]
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "t.xlsx"
            _write_rpa_xlsx(p, headers, data_rows)
            rows = load_rpa_sheet_hive_rows(p)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].table_description, "h1")

    def test_only_clickhouse_returns_empty_list(self) -> None:
        headers = [
            HEADER_TABLE_DESCRIPTION,
            HEADER_DDL,
            HEADER_STORAGE_PATH,
            HEADER_TABLE_TYPE,
        ]
        data_rows = [
            ["", "CREATE TABLE ck (i Int32);", "", "clickhouse"],
        ]
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "t.xlsx"
            _write_rpa_xlsx(p, headers, data_rows)
            rows = load_rpa_sheet_hive_rows(p)
        self.assertEqual(rows, [])

    def test_multiple_hive_rows_without_table_type_column(self) -> None:
        headers = [HEADER_TABLE_DESCRIPTION, HEADER_DDL, HEADER_STORAGE_PATH]
        data_rows = [
            ["d1", "sql1", "/p1"],
            ["d2", "sql2", "/p2"],
        ]
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "t.xlsx"
            _write_rpa_xlsx(p, headers, data_rows)
            rows = load_rpa_sheet_hive_rows(p)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0].table_description, "d1")
        self.assertEqual(rows[1].table_description, "d2")

    def test_unknown_table_type_skipped_with_warning(self) -> None:
        headers = [
            HEADER_TABLE_DESCRIPTION,
            HEADER_DDL,
            HEADER_STORAGE_PATH,
            HEADER_TABLE_TYPE,
        ]
        data_rows = [
            ["d", "sql", "/p", "starrocks"],
        ]
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "t.xlsx"
            _write_rpa_xlsx(p, headers, data_rows)
            with self.assertLogs("app.rpa_excel", level="WARNING") as cm:
                rows = load_rpa_sheet_hive_rows(p)
        self.assertEqual(rows, [])
        self.assertTrue(any("未知表类型" in m for m in cm.output))

    def test_skips_blank_middle_row(self) -> None:
        headers = [HEADER_TABLE_DESCRIPTION, HEADER_DDL, HEADER_STORAGE_PATH]
        data_rows = [
            ["d1", "sql1", "/p1"],
            ["", "", ""],
            ["d2", "sql2", "/p2"],
        ]
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "t.xlsx"
            _write_rpa_xlsx(p, headers, data_rows)
            rows = load_rpa_sheet_hive_rows(p)
        self.assertEqual(len(rows), 2)

    def test_empty_table_type_cell_treated_as_hive(self) -> None:
        headers = [
            HEADER_TABLE_DESCRIPTION,
            HEADER_DDL,
            HEADER_STORAGE_PATH,
            HEADER_TABLE_TYPE,
        ]
        data_rows = [["d", "sql", "/p", ""]]
        with tempfile.TemporaryDirectory() as td:
            p = Path(td) / "t.xlsx"
            _write_rpa_xlsx(p, headers, data_rows)
            rows = load_rpa_sheet_hive_rows(p)
        self.assertEqual(len(rows), 1)


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    unittest.main()
