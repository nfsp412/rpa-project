"""从 create_table_info.xlsx 的 `rpa` 工作表读取 RPA 表单字段。"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

SHEET_NAME = "rpa"
HEADER_TABLE_DESCRIPTION = "数据描述信息"
HEADER_DDL = "建表语句"
HEADER_STORAGE_PATH = "存储路径值"
HEADER_WAREHOUSE_LAYER = "数仓分层"
HEADER_TABLE_TYPE = "表类型"
HEADER_STATUS = "执行状态"

_DEFAULT_WAREHOUSE_LAYER = "ODS"


def default_excel_path() -> Path:
    """与 rpa-project 同级的 create-table-output/YYYYMMDD/create_table_info.xlsx（日期为运行当日）。"""
    rpa_project_root = Path(__file__).resolve().parent.parent
    workspace_root = rpa_project_root.parent
    day_dir = date.today().strftime("%Y%m%d")
    return workspace_root / "create-table-output" / day_dir / "create_table_info.xlsx"


@dataclass(frozen=True)
class RpaSheetRow:
    table_description: str
    ddl_sql: str
    storage_path: str
    warehouse_layer: str
    excel_row: int


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_warehouse_layer(raw: str) -> str:
    """与 DataWorks 下拉 option 文案对齐：空则默认 ODS，否则 strip 后转大写。"""
    s = raw.strip()
    if not s:
        return _DEFAULT_WAREHOUSE_LAYER
    return s.upper()


def _header_index(header_row: tuple[object, ...]) -> dict[str, int]:
    headers = [_cell_str(h) for h in header_row]
    idx: dict[str, int] = {}
    for i, name in enumerate(headers):
        if name and name not in idx:
            idx[name] = i
    return idx


def _col(data_row: tuple[object, ...], idx: dict[str, int], key: str) -> str:
    i = idx[key]
    if i >= len(data_row):
        return ""
    return _cell_str(data_row[i])


def _row_blank_for_rpa(data_row: tuple[object, ...], idx: dict[str, int]) -> bool:
    """数据描述与建表语句均为空则视为空行，跳过。"""
    return not _col(data_row, idx, HEADER_TABLE_DESCRIPTION) and not _col(
        data_row, idx, HEADER_DDL
    )


def _rpa_sheet_row_from_row(
    data_row: tuple[object, ...], idx: dict[str, int], excel_row: int
) -> RpaSheetRow:
    if HEADER_WAREHOUSE_LAYER in idx:
        layer_raw = _col(data_row, idx, HEADER_WAREHOUSE_LAYER)
    else:
        layer_raw = ""
    return RpaSheetRow(
        table_description=_col(data_row, idx, HEADER_TABLE_DESCRIPTION),
        ddl_sql=_col(data_row, idx, HEADER_DDL),
        storage_path=_col(data_row, idx, HEADER_STORAGE_PATH),
        warehouse_layer=_normalize_warehouse_layer(layer_raw),
        excel_row=excel_row,
    )


def load_rpa_sheet_hive_rows(xlsx_path: Path) -> list[RpaSheetRow]:
    """
    读取 `rpa` 表全部数据行，仅返回表类型为 Hive（列缺失或单元格空视为 Hive）的行。
    clickhouse 与未知非空表类型会跳过；未知类型会打 logging.warning。
    """
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"未找到 Excel 文件: {xlsx_path}")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(
                f"工作簿中不存在工作表 {SHEET_NAME!r}，现有: {wb.sheetnames}"
            )
        ws = wb[SHEET_NAME]
        row_iter = ws.iter_rows(min_row=1, values_only=True)
        try:
            header_row = next(row_iter)
        except StopIteration as e:
            raise ValueError(f"工作表 {SHEET_NAME!r} 为空") from e

        idx = _header_index(header_row)
        headers = [_cell_str(h) for h in header_row]
        missing = [
            h
            for h in (
                HEADER_TABLE_DESCRIPTION,
                HEADER_DDL,
                HEADER_STORAGE_PATH,
            )
            if h not in idx
        ]
        if missing:
            raise ValueError(
                f"表头缺少列: {missing}。当前表头: {[h for h in headers if h]}"
            )

        hive_rows: list[RpaSheetRow] = []
        substantive = 0

        for excel_row, data_row in enumerate(row_iter, start=2):
            if data_row is None:
                continue
            if _row_blank_for_rpa(data_row, idx):
                continue
            substantive += 1

            if HEADER_TABLE_TYPE in idx:
                raw_type = _col(data_row, idx, HEADER_TABLE_TYPE)
                t = raw_type.strip().lower()
            else:
                raw_type = ""
                t = ""

            if HEADER_STATUS in idx and _col(data_row, idx, HEADER_STATUS) == "成功":
                print(
                    f"[rpa] 第 {excel_row} 行已成功，跳过。",
                    flush=True,
                )
                continue

            if t == "clickhouse":
                print(
                    f"[rpa] 第 {excel_row} 行表类型为 clickhouse，已跳过。",
                    flush=True,
                )
                continue
            if t in ("", "hive"):
                hive_rows.append(_rpa_sheet_row_from_row(data_row, idx, excel_row))
                continue

            logger.warning(
                "第 %d 行未知表类型 %r，已跳过。",
                excel_row,
                raw_type,
            )

        if substantive == 0:
            raise ValueError(
                f"工作表 {SHEET_NAME!r} 没有有效数据行（至少需要一行含数据描述或建表语句）"
            )

        return hive_rows
    finally:
        wb.close()


def update_row_status(xlsx_path: Path, excel_row: int, status: str) -> None:
    """回写"执行状态"到指定行。若列不存在则自动追加表头。"""
    wb = load_workbook(xlsx_path)
    try:
        ws = wb[SHEET_NAME]
        header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]
        status_col: int | None = None
        for cell in header_cells:
            if _cell_str(cell.value) == HEADER_STATUS:
                status_col = cell.column
                break
        if status_col is None:
            status_col = len(header_cells) + 1
            ws.cell(row=1, column=status_col, value=HEADER_STATUS)
        ws.cell(row=excel_row, column=status_col, value=status)
        wb.save(xlsx_path)
    finally:
        wb.close()